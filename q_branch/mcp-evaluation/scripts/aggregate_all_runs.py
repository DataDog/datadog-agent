#!/usr/bin/env python3
"""
Aggregate evaluation results across all runs.

Creates a comprehensive analysis of all evaluation runs including:
- Average scores per scenario across modes
- Statistical analysis (mean, median, stddev)
- Comparison charts and detailed visualizations

Usage:
    python aggregate_all_runs.py [--results-dir DIRECTORY]

Reads from all run-* directories in specified directory and outputs aggregated_results.xlsx
"""

import argparse
import json
from pathlib import Path
from typing import Dict, List, Any
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
import sys

# Canonical scenario order (matches evaluate.py)
SCENARIO_ORDER = [
    # Easy
    "high-cpu-usage", "disk-space-full", "port-conflict",
    "zombie-processes", "dns-resolution-failure",
    # Medium
    "memory-leak", "connection-exhaustion", "log-rotation-failure",
    "swap-thrashing", "file-descriptor-leak",
    # Hard
    "tcp-close-wait", "io-wait", "context-switching-storm",
    "inode-exhaustion", "tcp-syn-flood"
]

# Scenario difficulty mapping
SCENARIO_DIFFICULTY = {
    # Easy
    "high-cpu-usage": "Easy",
    "disk-space-full": "Easy",
    "port-conflict": "Easy",
    "zombie-processes": "Easy",
    "dns-resolution-failure": "Easy",
    # Medium
    "memory-leak": "Medium",
    "connection-exhaustion": "Medium",
    "log-rotation-failure": "Medium",
    "swap-thrashing": "Medium",
    "file-descriptor-leak": "Medium",
    # Hard
    "tcp-close-wait": "Hard",
    "io-wait": "Hard",
    "context-switching-storm": "Hard",
    "inode-exhaustion": "Hard",
    "tcp-syn-flood": "Hard"
}


def load_all_runs(results_dir: Path, exclude_patterns: List[str] = None) -> List[Dict[str, Any]]:
    """Load all JSONL result files from all run directories

    Args:
        results_dir: Directory containing run-* subdirectories
        exclude_patterns: List of patterns to exclude (e.g., ['run-20260121*', 'run-20260122*'])
    """
    all_results = []
    run_dirs = sorted(results_dir.glob("run-*"))

    if not run_dirs:
        raise FileNotFoundError(f"No run directories found in {results_dir}")

    # Filter out excluded directories
    if exclude_patterns:
        excluded_dirs = set()
        for pattern in exclude_patterns:
            excluded_dirs.update(results_dir.glob(pattern))

        run_dirs = [d for d in run_dirs if d not in excluded_dirs]

        if excluded_dirs:
            print(f"Excluded {len(excluded_dirs)} directories:")
            for excluded_dir in sorted(excluded_dirs):
                print(f"  - {excluded_dir.name}")
            print()

    if not run_dirs:
        raise FileNotFoundError(f"No run directories found after exclusions in {results_dir}")

    print(f"Found {len(run_dirs)} run directories:")
    for run_dir in run_dirs:
        print(f"  - {run_dir.name}")

        for jsonl_file in run_dir.glob("evaluation-*.jsonl"):
            with open(jsonl_file, 'r') as f:
                for line in f:
                    if line.strip():
                        result = json.loads(line)
                        # Add run identifier
                        result['run'] = run_dir.name
                        all_results.append(result)

    print(f"\nLoaded {len(all_results)} total results across all runs\n")
    return all_results


def create_aggregated_stats(df: pd.DataFrame) -> pd.DataFrame:
    """Create aggregated statistics per mode and scenario across all runs"""
    # Filter to completed results only
    completed = df[df['status'] == 'completed'].copy()

    # Group by mode and scenario, calculate statistics
    stats = completed.groupby(['mode', 'scenario'])['overall_score'].agg([
        ('mean', 'mean'),
        ('median', 'median'),
        ('std', 'std'),
        ('min', 'min'),
        ('max', 'max'),
        ('count', 'count')
    ]).round(2)

    # Fill NaN std with 0 (happens when count=1)
    stats['std'] = stats['std'].fillna(0)

    # Reset index to make mode and scenario regular columns
    stats = stats.reset_index()

    # Sort by mode then scenario
    stats = stats.sort_values(['mode', 'scenario'])

    return stats


def create_mode_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Create summary statistics per mode across all scenarios and runs"""
    completed = df[df['status'] == 'completed'].copy()

    # Convert duration_ms to duration_sec
    completed['duration_sec'] = completed['duration_ms'] / 1000

    # Group by mode and calculate aggregates
    summary = completed.groupby('mode').agg({
        'overall_score': ['mean', 'min', 'max'],
        'duration_sec': 'mean',
        'turns': 'mean',
        'cost': ['mean', 'sum'],
        'scenario': 'count'  # Count of completed scenarios
    }).round(2)

    # Flatten multi-level columns
    summary.columns = ['avg_score', 'min_score', 'max_score', 'avg_duration_sec', 'avg_turns', 'avg_cost', 'total_cost', 'completed_count']

    # Add total count
    totals = df.groupby('mode').size()
    summary['total_scenarios'] = totals

    # Reset index to make mode a column
    summary = summary.reset_index()

    # Reorder columns (excluding completed_count)
    summary = summary[[
        'mode', 'total_scenarios',
        'avg_score', 'min_score', 'max_score',
        'avg_duration_sec', 'avg_turns', 'avg_cost', 'total_cost'
    ]]

    # Sort by avg_score in descending order
    summary = summary.sort_values('avg_score', ascending=False)

    return summary


def create_combined_comparison_chart(summary_sheet, stats_df: pd.DataFrame, output_file: Path):
    """
    Create a combined chart showing:
    1. Average scores per scenario across modes (bar chart)
    2. Error bars showing standard deviation
    """
    if stats_df.empty:
        print("No data to create chart")
        return

    # Pivot data: scenarios as rows, modes as columns, mean scores as values
    pivot_mean = stats_df.pivot(index='scenario', columns='mode', values='mean')
    pivot_std = stats_df.pivot(index='scenario', columns='mode', values='std')

    # Sort scenarios alphabetically
    pivot_mean = pivot_mean.sort_index()
    pivot_std = pivot_std.sort_index()

    # Get modes in consistent order
    mode_order = ['bash', 'safe-shell', 'tools', 'tools-safe-shell', 'tools-bash']
    available_modes = [m for m in mode_order if m in pivot_mean.columns]
    pivot_mean = pivot_mean[available_modes]
    pivot_std = pivot_std[available_modes]

    # Create figure
    fig, ax = plt.subplots(figsize=(18, 7))

    # Set up bar positions
    scenarios = pivot_mean.index.tolist()
    x = np.arange(len(scenarios))
    width = 0.15  # Narrower bars for 5 modes

    # Colors for each mode
    colors = {
        'bash': '#1f77b4',              # Blue
        'safe-shell': '#ff7f0e',        # Orange
        'tools': '#2ca02c',             # Green
        'tools-safe-shell': '#d62728',  # Red
        'tools-bash': '#9467bd'         # Purple
    }

    # Create bars for each mode with error bars
    for i, mode in enumerate(available_modes):
        offset = width * (i - len(available_modes)/2 + 0.5)
        means = pivot_mean[mode].values
        stds = pivot_std[mode].values

        bars = ax.bar(x + offset, means, width,
                     label=mode, color=colors.get(mode, '#333333'),
                     edgecolor='black', linewidth=0.5,
                     yerr=stds, capsize=3, error_kw={'linewidth': 1, 'elinewidth': 1})

    # Customize chart
    ax.set_xlabel('Scenario', fontsize=14, fontweight='bold')
    ax.set_ylabel('Score (0-100)', fontsize=14, fontweight='bold')
    ax.set_title('Average Diagnostic Scores Across All Runs (with Std Dev)',
                fontsize=16, fontweight='bold', pad=20)
    ax.set_xticks(x)
    ax.set_xticklabels(scenarios, rotation=45, ha='right', fontsize=10)
    ax.set_ylim(0, 105)
    ax.set_yticks(range(0, 101, 10))
    ax.legend(title='Mode', loc='upper left', bbox_to_anchor=(1.02, 1), fontsize=12)
    ax.grid(axis='y', alpha=0.3, linestyle='--')

    # Tight layout
    plt.tight_layout()

    # Save figure
    chart_file = output_file.parent / "aggregated_score_comparison.png"
    plt.savefig(chart_file, dpi=100, bbox_inches='tight')
    plt.close()

    print(f"Aggregated score comparison chart saved to {chart_file}")

    # Embed in Excel
    img = Image(str(chart_file))
    img.width = img.width * 0.75
    img.height = img.height * 0.75
    summary_sheet.add_image(img, 'A15')

    print("Embedded aggregated chart in Summary sheet")


def create_statistics_chart(summary_sheet, stats_df: pd.DataFrame, output_file: Path):
    """
    Create a detailed statistics visualization showing mean, median, and range
    for each mode across scenarios.
    """
    if stats_df.empty:
        print("No data to create statistics chart")
        return

    # Get modes
    mode_order = ['bash', 'safe-shell', 'tools', 'tools-safe-shell', 'tools-bash']
    available_modes = [m for m in mode_order if m in stats_df['mode'].unique()]

    # Create subplots - one for each mode
    fig, axes = plt.subplots(1, len(available_modes), figsize=(18, 6), sharey=True)

    if len(available_modes) == 1:
        axes = [axes]

    colors = {
        'bash': '#1f77b4',              # Blue
        'safe-shell': '#ff7f0e',        # Orange
        'tools': '#2ca02c',             # Green
        'tools-safe-shell': '#d62728',  # Red
        'tools-bash': '#9467bd'         # Purple
    }

    for idx, mode in enumerate(available_modes):
        ax = axes[idx]
        mode_data = stats_df[stats_df['mode'] == mode].sort_values('mean')

        scenarios = mode_data['scenario'].tolist()
        means = mode_data['mean'].values
        medians = mode_data['median'].values
        stds = mode_data['std'].values
        mins = mode_data['min'].values
        maxs = mode_data['max'].values

        y = np.arange(len(scenarios))

        # Plot range (min to max) as horizontal lines
        for i in range(len(scenarios)):
            ax.plot([mins[i], maxs[i]], [y[i], y[i]],
                   color=colors[mode], alpha=0.3, linewidth=8)

        # Plot mean as filled circles
        ax.scatter(means, y, s=100, color=colors[mode],
                  edgecolor='black', linewidth=1.5, zorder=3, label='Mean')

        # Plot median as X markers
        ax.scatter(medians, y, s=100, marker='x', color='black',
                  linewidth=2, zorder=4, label='Median')

        # Customize subplot
        ax.set_title(f'{mode.upper()}', fontsize=14, fontweight='bold',
                    color=colors[mode])
        ax.set_xlabel('Score', fontsize=12)
        ax.set_yticks(y)
        # Add difficulty level in parentheses after scenario name
        scenario_labels = [f"{s} ({SCENARIO_DIFFICULTY.get(s, 'Unknown')})" for s in scenarios]
        ax.set_yticklabels(scenario_labels, fontsize=9)
        ax.set_xlim(-5, 105)
        # Add both horizontal and vertical grid lines
        ax.grid(axis='x', alpha=0.3, linestyle='--')
        ax.grid(axis='y', alpha=0.2, linestyle='-', linewidth=0.5)

        # Add legend to first subplot only
        if idx == 0:
            ax.set_ylabel('Scenario', fontsize=12, fontweight='bold')
            # Create custom legend
            from matplotlib.lines import Line2D
            legend_elements = [
                Line2D([0], [0], marker='o', color='w', markerfacecolor='gray',
                      markersize=10, markeredgecolor='black', linewidth=1.5, label='Mean'),
                Line2D([0], [0], marker='x', color='w', markerfacecolor='black',
                      markersize=10, markeredgecolor='black', linewidth=2, label='Median'),
                Line2D([0], [0], color='gray', alpha=0.5, linewidth=8, label='Min-Max Range')
            ]
            ax.legend(handles=legend_elements, loc='lower right', fontsize=9)

    # Overall title
    fig.suptitle('Statistical Summary: Mean, Median, and Range by Mode',
                fontsize=16, fontweight='bold', y=0.99)

    # Tight layout with extra top padding
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    # Save figure
    chart_file = output_file.parent / "statistics_summary.png"
    plt.savefig(chart_file, dpi=100, bbox_inches='tight')
    plt.close()

    print(f"Statistics summary chart saved to {chart_file}")

    # Embed in Excel (below the first chart)
    img = Image(str(chart_file))
    img.width = img.width * 0.75
    img.height = img.height * 0.75
    summary_sheet.add_image(img, 'A50')

    print("Embedded statistics chart in Summary sheet")


def write_aggregated_excel(results: List[Dict[str, Any]], output_file: Path):
    """Write aggregated results to Excel with multiple sheets and charts"""
    if not results:
        print("No results to write")
        return

    # Create DataFrame
    records = []
    for r in results:
        records.append({
            'run': r.get('run'),
            'mode': r.get('mode'),
            'scenario': r.get('scenario'),
            'status': r.get('status'),
            'overall_score': r.get('score', {}).get('overall_score'),
            'duration_ms': r.get('duration_ms'),
            'turns': r.get('turns'),
            'cost': r.get('cost'),
        })

    df = pd.DataFrame(records)

    # Calculate aggregated statistics
    print("\nCalculating aggregated statistics...")
    stats_df = create_aggregated_stats(df)
    mode_summary = create_mode_summary(df)

    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write mode summary sheet
        print("Creating Mode Summary sheet...")
        mode_summary.to_excel(writer, sheet_name='Mode Summary', index=False)

        # Write detailed statistics sheet with separate tables per mode
        print("Creating Detailed Statistics sheet...")
        stats_sheet = writer.book.create_sheet('Detailed Statistics')

        current_row = 1
        mode_order = ['bash', 'safe-shell', 'tools', 'tools-safe-shell', 'tools-bash']
        available_modes = [m for m in mode_order if m in stats_df['mode'].unique()]

        for mode in available_modes:
            mode_stats = stats_df[stats_df['mode'] == mode].copy()
            mode_stats = mode_stats.drop('mode', axis=1)  # Remove mode column as it's in the header

            # Write mode header
            header_cell = stats_sheet.cell(row=current_row, column=1, value=f'{mode.upper()} Statistics')
            header_cell.font = Font(bold=True, size=14)
            current_row += 1

            # Write column headers
            for col_idx, col_name in enumerate(mode_stats.columns, start=1):
                col_header_cell = stats_sheet.cell(row=current_row, column=col_idx, value=col_name)
                col_header_cell.font = Font(bold=True)
            current_row += 1

            # Write data rows
            for _, row in mode_stats.iterrows():
                for col_idx, value in enumerate(row, start=1):
                    stats_sheet.cell(row=current_row, column=col_idx, value=value)
                current_row += 1

            # Add spacing between mode tables
            current_row += 2

        # Write individual sheets per mode with all runs
        for mode in sorted(df['mode'].unique()):
            print(f"Creating {mode} All Runs sheet...")
            mode_df = df[df['mode'] == mode].copy()
            mode_df = mode_df.sort_values(['scenario', 'run']).reset_index(drop=True)
            mode_df.to_excel(writer, sheet_name=f'{mode} All Runs', index=False)

        # Write raw data sheet
        print("Creating Raw Data sheet...")
        df.to_excel(writer, sheet_name='Raw Data', index=False)

        # Auto-adjust column widths and format currency columns
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            df_for_sheet = None

            # Determine which dataframe is in this sheet
            if sheet_name == 'Mode Summary':
                df_for_sheet = mode_summary
            elif sheet_name == 'Raw Data':
                df_for_sheet = df
            elif sheet_name.endswith(' All Runs'):
                mode_name = sheet_name.replace(' All Runs', '')
                df_for_sheet = df[df['mode'] == mode_name]

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format cost columns as currency
            if df_for_sheet is not None:
                for idx, col in enumerate(df_for_sheet.columns):
                    if 'cost' in col.lower():
                        col_letter = chr(65 + idx)
                        for row in range(2, len(df_for_sheet) + 2):  # Start from row 2 (skip header)
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = '$#,##0.00'

        # Add charts to Mode Summary sheet
        summary_sheet = writer.sheets['Mode Summary']

        print("\nCreating visualizations...")
        create_combined_comparison_chart(summary_sheet, stats_df, output_file)
        create_statistics_chart(summary_sheet, stats_df, output_file)

    print(f"\nAggregated results written to {output_file}")

    # Print summary to console
    print("\n" + "="*80)
    print("MODE SUMMARY (Across All Runs)")
    print("="*80)
    print(mode_summary.to_string(index=False))
    print("="*80)

    # Print top/bottom performers
    print("\n" + "="*80)
    print("TOP 5 PERFORMING SCENARIOS (by mean score)")
    print("="*80)
    top_5 = stats_df.nlargest(5, 'mean')[['mode', 'scenario', 'mean', 'std', 'count']]
    print(top_5.to_string(index=False))

    print("\n" + "="*80)
    print("BOTTOM 5 PERFORMING SCENARIOS (by mean score)")
    print("="*80)
    bottom_5 = stats_df.nsmallest(5, 'mean')[['mode', 'scenario', 'mean', 'std', 'count']]
    print(bottom_5.to_string(index=False))
    print("="*80)


def parse_args():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description="Aggregate evaluation results across all runs",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Aggregate all runs in default results/ directory
  python aggregate_all_runs.py

  # Aggregate runs in baseline directory
  python aggregate_all_runs.py --results-dir results/baseline

  # Aggregate runs in hybrid-modes directory
  python aggregate_all_runs.py --results-dir results/hybrid-modes

  # Exclude specific runs by pattern
  python aggregate_all_runs.py --exclude "run-20260121*" --exclude "run-20260122*"

  # Combine directory selection with exclusions
  python aggregate_all_runs.py --results-dir results/baseline --exclude "run-20260121_133310"
        """
    )

    parser.add_argument(
        "--results-dir",
        type=Path,
        default=None,
        help="Directory containing run-* subdirectories (default: results/)"
    )

    parser.add_argument(
        "--exclude",
        action="append",
        default=[],
        help="Exclude directories matching pattern (can be specified multiple times)"
    )

    return parser.parse_args()


def main():
    args = parse_args()

    # Determine paths
    script_dir = Path(__file__).parent
    mcp_eval_dir = script_dir.parent

    # Use specified directory or default to results/
    if args.results_dir:
        results_base_dir = args.results_dir
        if not results_base_dir.is_absolute():
            # Make relative paths relative to mcp_eval_dir
            results_base_dir = mcp_eval_dir / results_base_dir
    else:
        results_base_dir = mcp_eval_dir / "results"

    if not results_base_dir.exists():
        print(f"Error: Results directory does not exist: {results_base_dir}")
        return 1

    # Output file goes in the specified results directory
    output_file = results_base_dir / "aggregated_results.xlsx"

    print(f"Results directory: {results_base_dir}")
    print(f"Output file: {output_file}\n")

    # Load all runs
    try:
        results = load_all_runs(results_base_dir, exclude_patterns=args.exclude)
    except Exception as e:
        print(f"Error loading results: {e}")
        return 1

    if not results:
        print("No results found")
        return 1

    # Write aggregated Excel with charts
    try:
        write_aggregated_excel(results, output_file)
    except Exception as e:
        print(f"Error writing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return 1

    print(f"\n✓ Aggregation complete!")
    print(f"\nTo use in Google Sheets:")
    print(f"1. Open Google Sheets")
    print(f"2. File > Import > Upload")
    print(f"3. Select: {output_file}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
